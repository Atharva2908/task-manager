from fastapi import APIRouter, Query, HTTPException, Response
from typing import Literal
import io
import csv
from datetime import datetime
from bson import ObjectId

# Import your database connection function
from app.database import get_db

router = APIRouter()

async def get_tasks_from_db():
    """Fetch all tasks from MongoDB"""
    db = get_db()
    tasks_collection = db["tasks"]
    users_collection = db["users"]
    
    tasks = await tasks_collection.find({"is_deleted": False}).to_list(length=1000)
    
    # Convert ObjectId to string and enrich with user names
    for task in tasks:
        task['_id'] = str(task['_id'])
        
        # Attempt to get assigned user's full name if available
        if task.get('assigned_to'):
            try:
                user = await users_collection.find_one({"_id": ObjectId(task['assigned_to'])})
                if user:
                    task['assigned_to'] = f"{user.get('first_name', '')} {user.get('last_name', '')}".strip()
            except Exception:
                # Fail silently if error occurs
                pass
    
    return tasks

@router.get("/export")
async def export_report(
    format: Literal['csv', 'excel', 'pdf'] = Query(..., description="Export format")
):
    """Public endpoint to export tasks report in CSV, Excel, or PDF format (no authentication)."""
    try:
        tasks = await get_tasks_from_db()
        if not tasks:
            tasks = []
        
        if format == 'csv':
            return export_csv(tasks)
        elif format == 'excel':
            return export_excel(tasks)
        elif format == 'pdf':
            return export_pdf(tasks)
        else:
            raise HTTPException(status_code=400, detail="Invalid export format")
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")

def export_csv(tasks):
    output = io.StringIO()
    writer = csv.writer(output)
    
    writer.writerow(['Title', 'Status', 'Priority', 'Due Date', 'Assigned To', 'Created At'])
    
    for task in tasks:
        writer.writerow([
            task.get('title', ''),
            task.get('status', ''),
            task.get('priority', ''),
            task.get('due_date', ''),
            task.get('assigned_to', ''),
            task.get('created_at', '')
        ])
    
    headers = {
        'Content-Disposition': f'attachment; filename="tasks_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    }
    
    return Response(
        content=output.getvalue(),
        media_type='text/csv',
        headers=headers
    )

def export_excel(tasks):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed. Run: pip install openpyxl")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Tasks Report"
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    headers = ['Title', 'Status', 'Priority', 'Due Date', 'Assigned To', 'Created At']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for row_num, task in enumerate(tasks, 2):
        ws.cell(row=row_num, column=1, value=task.get('title', ''))
        ws.cell(row=row_num, column=2, value=task.get('status', ''))
        ws.cell(row=row_num, column=3, value=task.get('priority', ''))
        ws.cell(row=row_num, column=4, value=task.get('due_date', ''))
        ws.cell(row=row_num, column=5, value=task.get('assigned_to', ''))
        ws.cell(row=row_num, column=6, value=task.get('created_at', ''))
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    headers = {
        'Content-Disposition': f'attachment; filename="tasks_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    }
    
    return Response(
        content=buffer.getvalue(),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers=headers
    )

def export_pdf(tasks):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    except ImportError:
        raise HTTPException(status_code=500, detail="reportlab not installed. Run: pip install reportlab")
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    styles = getSampleStyleSheet()
    title = Paragraph(f"<b>Tasks Report</b><br/>{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 20))
    
    data = [['Title', 'Status', 'Priority', 'Due Date', 'Assigned To']]
    for task in tasks[:50]:
        data.append([
            task.get('title', '')[:30],
            task.get('status', ''),
            task.get('priority', ''),
            task.get('due_date', '')[:10] if task.get('due_date') else '',
            task.get('assigned_to', '')[:20]
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))
    
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    
    headers = {
        'Content-Disposition': f'attachment; filename="tasks_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    }
    
    return Response(
        content=buffer.getvalue(),
        media_type='application/pdf',
        headers=headers
    )
